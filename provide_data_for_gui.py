from pymysql import *
import datetime
import config as con
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import Workbook


def Build_table(head, path):
    # 在内存中创建一个workbook对象，并且至少创建一个worksheet
    wb = Workbook()

    ws = wb.active  # 获取当前活跃的worksheet，默认是第一个sheet

    for i in range(1, len(head) + 1):
        ws.cell(1, i).value = head[i - 1]

    wb.save(filename=path)


def get_time_now():
    time_now = datetime.datetime.now().strftime('%Y%m%d%H')
    return time_now


def Connect():  # 连接数据库
    db = connect(host='localhost', user="root", password=con.password, db="Program", charset="utf8")
    cursor = db.cursor()
    return db, cursor


def get_by_hour(time):  # 具体一个小时的数据
    '''
    :param time: 传入数据为字符串形式 'date + hour'
    :return:
    '''
    db, cursor = Connect()
    date = time[:8]
    hour = int(time[8:])
    sql = "SELECT * FROM all_data WHERE date = '%s' and time = '%s'" % (date, hour)
    cursor.execute(sql)
    result = cursor.fetchall()[0]
    result = result[1:]
    state = result[-1]     # 取现在的状态看是否已经是删除的数据
    if state == 0:
        return False
    result = result[:-1]   # 去掉末尾
    name = []
    for x in cursor.description[1:]:
        name.append(x[0])
    name = name[:-1]
    # print(name)
    # print(result)
    return name, result


def get_by_day(time):  # 数据库中有ID栏
    '''
    :param time:'date'
    :return:
    '''
    db, cursor = Connect()
    sql = "SELECT * FROM all_data WHERE date = '%s' and state = 1" % (time)
    cursor.execute(sql)
    data = cursor.fetchall()
    result = []
    for x in data:
        result.append(x[1:])
    name = []
    for x in cursor.description[1:]:
        name.append(x[0])
    name = name[:-1]
    result = list(map(list, zip(*result)))
    result = result[2:]
    result = result[:-1]
    return name, result


def get_by_fragment_(time_now=get_time_now(), number=72):
    '''
    :param time_now: 'date + hour'
            number : fragment的长度
    :return:
    '''
    db, cursor = Connect()
    '''
    date = time_now[:8]
    hour = int(time_now[8:]) - 1
    '''
    '''
    sql = "SELECT * FROM all_data WHERE date = '%s' and time = '%s' and state = 1" % (date, hour)
    result = []
    if cursor.execute(sql):
        name = []
        for x in cursor.description[1:]:
            name.append(x[0])
        data_now = cursor.fetchall()[0]
        id = data_now[0]
        # print(id)
        result.append(data_now[1:])  # 添加第一个元素给result
        i = 1
    else:       #若不能
        sql = "SELECT ID FROM all_data where state = 1"
        cursor.execute(sql)
        id = cursor.fetchall()[-1][0]
        i = 0
        name = get_table_name()
        print(id)
    while i != number - 1 and id - i != 0:
        sql = "SELECT * FROM all_data WHERE ID = '%d' and state = 1" % (id - i)
        cursor.execute(sql)
        result.append(cursor.fetchall()[0][1:])
        # print(result)
        i += 1
        # print(i)
    length = len(result)
    T_result = [[] for i in range(length)]
    for i in range(length):
        T_result[length - i - 1] = result[i]
    '''
    sql = 'SELECT * FROM all_data WHERE state=1 ORDER BY id DESC LIMIT 72'
    try:
        cursor.execute(sql)
        result = cursor.fetchall()
        result = [x[1:-1] for x in result]
        length = len(result)
        T_result = [[] for i in range(length)]
        for i in range(length):
            T_result[length - i - 1] = result[i]
        name = get_table_name()
        #print(T_result)
        return T_result, name
    except Exception as msg:
        print(msg)
        return False, False
    # return T_result, name
    # print(result)


def get_by_fragment():
    '''
    以实时时间为节点，获取前72小时的数据
    :return:
    '''
    time_now = get_time_now()
    T_result, name = get_by_fragment_(time_now)
    return T_result, name


def update_data(data):
    db, cursor = Connect()
    try:
        for x in data:
            date = x[0]
            hour = x[1]
            name, raw_data = get_by_hour(str(date) + str(hour))
            for i in range(2, len(x)):
                if x[i] != raw_data[i]:
                    sql = "UPDATE all_data SET %s = '%lf' WHERE date = '%s' and time = '%s'" \
                          % (name[i], x[i], date, hour)
                    cursor.execute(sql)
                    db.commit()
        return True
    except Exception as msg:
        print(msg)
        return False


def delete_data(data):
    db, cursor = Connect()
    try:
        for x in data:
            date = x[0]
            hour = x[1]
            sql = "UPDATE all_data SET state = 0 where date=%s and time=%s" % (date, hour)
            cursor.execute(sql)
            db.commit()

        return True
    except Exception as msg:
        print(msg)
        return False


def recover_data(date):
    '''
    :param date: 提供需要恢复的数据的日期和时间信息:date+hour
    :return: True or False
    '''


def save_data(data):
    db, cursor = Connect()
    try:
        sql1 = "SELECT * FROM all_data WHERE ID = 1"
        cursor.execute(sql1)
        information = cursor.description
        name = []
        for x in information:
            name.append(x[0])
        Name = name[1:]
        # print(name)

        sql = "SELECT ID FROM all_data"
        cursor.execute(sql)
        ids = cursor.fetchall()
        # print(ids)
        last_id = ids[-1][0]
        for i in range(len(data)):
            # Data = Name + data[i]
            # print(data[i])
            sql = "INSERT INTO all_data(ID, date, time) VALUE('%d', '%s', '%s')" % (last_id + 1, data[i][0], data[i][1])
            last_id += 1
            cursor.execute(sql)
            db.commit()
            for j in range(2, len(data[0])):
                if data[i][j] == None or data[i][j] == '':
                    data[i][j] = 0
                sql = "UPDATE all_data SET %s = '%lf' WHERE date = '%s' and time = '%s' and ID = '%d'" \
                      % (Name[j], data[i][j], data[i][0], data[i][1], last_id)
                cursor.execute(sql)
                db.commit()
            '''
            sql = "INSERT INTO all_data(%s for i in range(len(Name))) VALUE('%s' for i in range(len(data[0])))"\
                  % (Data[j] for j in range(len(Data)))
                  '''
        print('OK')
        return True
    except Exception as msg:
        print(msg)
        return False


def get_table_name():
    db, cursor = Connect()
    sql = "SELECT * FROM all_data WHERE ID = 1"
    cursor.execute(sql)
    result = cursor.description
    name = [x[0] for x in result]
    # print(name[1:])
    return name[1:-1]


def get_all_date():
    db, cursor = Connect()
    sql = "SELECT date, time FROM all_data WHERE state = 1"
    cursor.execute(sql)
    result = cursor.fetchall()
    return result


def get_target_data(time):
    db, cursor = Connect()
    date = time[:8]
    hour = time[8:]
    sql = "SELECT * FROM target_data WHERE date = '%s' and time = '%s'" % (date, hour)
    if cursor.execute(sql):
        name = cursor.description
        name = [x[0] for x in name[1:]]
        result = cursor.fetchall()[0][1:]
        return name, result
    else:
        print("There is no data")
        return False


def reshape_data(source_filename, target_filename='file_model\\raw_model.xlsx'):
    '''
    :param source_filename:
    :param target_filename:
    :return: True of False
    '''
    source_readfile = xlrd.open_workbook(source_filename)
    sourcesheet = source_readfile.sheet_by_index(0)
    source_table_name = sourcesheet.row_values(0, 0)  # 获取用户文件的表头

    standard_file = xlrd.open_workbook(target_filename)
    standardsheet = standard_file.sheet_by_name('Sheet1')
    standard_table_name = standardsheet.row_values(0, 0)

    new_table_value = [[] for i in range(len(standard_table_name))]

    for i in range(len(standard_table_name)):
        lable = 0
        for j in range(len(source_table_name)):
            if standard_table_name[i] == source_table_name[j]:
                table_value = sourcesheet.col_values(j, 1)
                new_table_value[i] = table_value
                lable = 1
                break
        if lable == 0:
            new_table_value[i] = ['' for x in range(len(standardsheet.col_values(0, 1)))]

    new_file = xlwt.Workbook()
    newsheet = new_file.add_sheet('Sheet1')
    for i in range(len(standard_table_name)):
        newsheet.write(0, i, standard_table_name[i])
    for i in range(len(standard_table_name)):
        for j in range(len(new_table_value[i])):
            newsheet.write(j + 1, i, new_table_value[i][j])

    length = len('model\\model')
    new_filename = 'data' + target_filename[length:]
    print(new_filename)
    if new_file.save('Datafile_From_C\\' + new_filename):
        return True
    else:
        return False


def change_data_model(Username, name):
    '''
    :param Username: 更改该模板的用户信息
    :param name: 更改后的表头名称
    :return: 新模板的文件路径 or False
    '''
    oldfile = xlrd.open_workbook('file_model\\raw_model.xlsx')
    newfile = copy(oldfile)
    new_filename = 'file_model\\model_for_' + Username + '.xls'
    new_sheet = newfile.get_sheet(0)
    try:
        length = len(name)
        for i in range(length):
            new_sheet.write(0, i, name[i])
        newfile.save(new_filename)
        return new_filename
    except Exception:
        return False


def Read_file(path):  # 读取文件名，返回文件的所有数据
    readfile = xlrd.open_workbook(path)
    value = []
    read_sheet = readfile.sheet_by_name('Sheet1')
    table_name = read_sheet.row_values(0, 0)
    for i in range(len(table_name)):
        x = read_sheet.col_values(i, 1)
        value.append(x)
    return table_name, value


def operation_record(data):
    try:
        db, cursor = Connect()
        time_now = datetime.datetime.now().strftime('%Y%m%d-%H:%M:%S')
        # print(time_now)
        sql = "INSERT INTO operation_record(time, username, operation) VALUE('%s', '%s', '%s')" \
              % (time_now, data[0], data[1])
        print('laalalala')
        cursor.execute(sql)
        db.commit()
        return True
    except Exception as msg:
        print('Operation_record_error:'+msg)
        return False


def Data_Output_from_Database():        #  从数据库中导出数据
    try:
        db, cursor = Connect()
        sql = "SELECT * FROM all_data WHERE state = 1"
        cursor.execute(sql)
        new_file = xlwt.Workbook()
        newsheet = new_file.add_sheet('Sheet1')
        time = datetime.datetime.now().strftime('%Y%m%d%H%M')
        new_file.save('E:\\Data_Output\\all_data-' + time + '.xls')
        record = cursor.fetchall()
        table_name = get_table_name()
        for i in range(len(table_name)):
            newsheet.write(0, i, table_name[i])     #  写入表头
        for i in range(len(record)):
            for j in range(len(record[i])-2):   # 此处-2是不记录state和id
                newsheet.write(i+1, j, record[i][j+1])       # 因为是从第二行开始录入数据，所以i+1
        # print(time)
        new_file.save('E:\\Data_Output\\all_data-'+time+'.xls')
        return True
    except Exception as msg:
        print(msg)
        return False

def get_chinese(table_name):
    '''
    :param table_name:
    :return:
    '''
    db, cursor = Connect()
    cursor = db.cursor()
    sql = "SELECT show_name FROM nameToChinese WHERE table_name='%s';" % (table_name)
    #print(sql)
    cursor.execute(sql)
    result = cursor.fetchone()
    if result==None:
        return table_name
    else:
        return result[0]


if __name__ == "__main__":
    # Data_Output_from_Database()
    get_by_fragment_()
    '''
    old_table_name = get_table_name()
    old_table_name.append('zhu')
    model_name = change_data_model('moujun', old_table_name)
    reshape_data('source.xlsx', model_name)
    # print(get_target_data('2017012310'))
    # get_all_date()
    # get_table_name()
    # print(get_time_now())
    # get_by_fragment()
    
    name1, data1 = get_by_hour('2017031507')
    name2, data2 = get_by_hour('2017031508')
    data1 = list(data1)
    data2 = list(data2)
    # print([data1, data2])
    # data[3] = 400

    if save_data([data1, data2]):
        print('OK')
    else:
        print('FALSE')
'''

